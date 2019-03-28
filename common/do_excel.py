# -*- coding: utf-8 -*-
# @Time  : 2019/3/23 11:04
# @Author: Tansy_Xiaoming
# @Email : 279244228@qq.com
# @File  : do_excel.py
from openpyxl import load_workbook
from openpyxl import Workbook
from class_0313_api_practice_4.common import common_path
from class_0313_api_practice_4.common.read_config import ReadConfig


class DoExcel:
    '''根据配置文件，对excel中数据进行全部读取、单个读取和写入'''
    flag = ReadConfig().get_str('TestSet','RunCase') # 从配置文件中读取执行用例的配置

    def __init__(self,sheetname,filename=common_path.excel_path):
        self.filename = filename
        self.sheetname = sheetname

    def read_all_data(self):
        '''根据配置文件中定义所有跑的用例，输出用例数据'''
        wb = load_workbook(self.filename)
        sheet = wb[self.sheetname]
        file_data = []
        for i in range(2, sheet.max_row+1):
            row_data = {}
            row_data['caseId'] = sheet.cell(i,1).value
            row_data['description'] = sheet.cell(i,2).value
            row_data['method'] = sheet.cell(i,3).value
            row_data['params'] = sheet.cell(i,4).value
            row_data['expectedResult'] = sheet.cell(i,5).value
            row_data['sql'] = sheet.cell(i,6).value
            if self.flag.upper() == 'ALL':  # 如果配置文件是执行全部用例，则读取的每行数据都添加到最终输出数据中
                file_data.append(row_data)
            else:
                flag = eval(self.flag)  # 用例的配置文件格式为字典 嵌套列表
                if self.sheetname in flag.keys():   # 表单名即为模块名，判断是否有设置此模块的用例
                    if row_data['caseId'] in flag[self.sheetname]: # 判断当前读取出来的用例是否在配置文件中
                        file_data.append(row_data)
        wb.close()
        return file_data

    def read_one_data(self,row,column):
        '''读取某个单元格的值并返回'''
        wb = load_workbook(self.filename)
        sheet = wb[self.sheetname]
        res = sheet.cell(row,column).value
        wb.close()
        return res

    def update_excel(self,row,column,value):
        '''写入某个单元格的值'''
        wb = load_workbook(self.filename)
        sheet = wb[self.sheetname]
        try:
            sheet.cell(row,column).value = value
            wb.save(self.filename)
            wb.close()
        except Exception as e:
            print('写入数据错误，错误为{}'.format(e))


if __name__ == '__main__':
    w = DoExcel('Login')
    print(w.read_all_data())
    # print(w.read_one_data(1,2))
    # print(type(w.read_one_data(1,2)))